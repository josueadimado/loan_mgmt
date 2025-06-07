# investors/views.py
from django.shortcuts import render, get_object_or_404
from django.urls import reverse_lazy
from django.views.generic import ListView, UpdateView, DetailView, View
from django.http import HttpResponse
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from datetime import datetime
from decimal import Decimal
from .models import Investor, InvestorTransaction
from .forms import InvestorForm, InvestorTransactionForm

class InvestorListView(LoginRequiredMixin, ListView):
    model = Investor
    template_name = 'investors/investors_list.html'
    context_object_name = 'investors'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['count'] = Investor.objects.count()
        # Provide a form for adding a new investor
        ctx['form'] = InvestorForm(prefix='investor', user=self.request.user)
        # Default to not showing the form
        ctx['show_form'] = self.request.GET.get('show_form', 'false') == 'true'
        print(f"InvestorListView: Rendering template {self.template_name}")
        return ctx

    def post(self, request, *args, **kwargs):
        form = InvestorForm(request.POST, request.FILES, prefix='investor', user=request.user)
        if form.is_valid():
            investor = form.save()
            # Handle initial investment
            initial_investment = form.cleaned_data.get('initial_investment')
            if initial_investment and initial_investment > 0:
                transaction = InvestorTransaction.objects.create(
                    investor=investor,
                    transaction_type='deposit',
                    amount=initial_investment,
                    created_by=self.request.user,
                )
                investor.funds_available += initial_investment
                investor.save()
                print(f"InvestorListView: Created initial deposit transaction: {transaction}")
                messages.success(self.request, "Investor created successfully with initial investment.")
            else:
                messages.success(self.request, "Investor created successfully.")
            return self.redirect_with_query(show_form='false')
        else:
            messages.error(self.request, "Please correct the errors below.")
            print(f"InvestorListView: Form invalid. Errors: {form.errors}")
            self.object_list = self.get_queryset()
            context = self.get_context_data()
            context['form'] = form
            context['show_form'] = True
            return self.render_to_response(context)

    def redirect_with_query(self, **kwargs):
        """Helper method to redirect with query parameters."""
        url = reverse_lazy('investors:investor-list')
        from urllib.parse import urlencode
        query = urlencode(kwargs)
        return HttpResponse(status=302, headers={'Location': f"{url}?{query}"})

class InvestorUpdateView(LoginRequiredMixin, UpdateView):
    model = Investor
    form_class = InvestorForm
    template_name = 'investors/investors_form.html'
    success_url = reverse_lazy('investors:investor-list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_queryset(self):
        return Investor.objects.filter(created_by=self.request.user)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['investor'] = self.object
        context['transactions'] = self.object.transactions.all().order_by('-date')
        print(f"InvestorUpdateView: Rendering template {self.template_name}")
        return context

    def form_valid(self, form):
        if not form.has_changed():
            messages.info(self.request, "No changes detected.")
            return super().form_valid(form)

        changes = []
        for field in form.changed_data:
            submitted_value = form.cleaned_data.get(field)
            existing_value = getattr(self.object, field)
            changes.append(f"{field}: submitted={submitted_value}, existing={existing_value}")
        if changes:
            print("InvestorUpdateView: Changes detected:", changes)

        response = super().form_valid(form)
        messages.success(self.request, "Investor updated successfully.")
        return response

    def form_invalid(self, form):
        messages.error(self.request, "Please correct the errors below.")
        print(f"InvestorUpdateView: Form invalid. Errors: {form.errors}")
        print(f"InvestorUpdateView: POST data: {self.request.POST}")
        return super().form_invalid(form)

class InvestorDetailView(LoginRequiredMixin, DetailView):
    model = Investor
    template_name = 'investors/investors_detail.html'
    context_object_name = 'investor'

    def get_queryset(self):
        return Investor.objects.filter(created_by=self.request.user)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['transactions'] = self.object.transactions.all().order_by('-date')
        # Add top-up form
        context['topup_form'] = InvestorTransactionForm(user=self.request.user, investor=self.object)
        print(f"InvestorDetailView: Rendering template {self.template_name}")
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = InvestorTransactionForm(request.POST, user=request.user, investor=self.object)
        if form.is_valid():
            form.save()
            messages.success(request, "Top-up added successfully.")
            return redirect('investors:investor-detail', pk=self.object.id)
        else:
            messages.error(request, "Please correct the errors below.")
            context = self.get_context_data()
            context['topup_form'] = form
            return self.render_to_response(context)

class InvestorExportView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        import datetime

        qs = Investor.objects.filter(created_by=self.request.user).order_by('first_name', 'last_name')
        wb = Workbook()
        ws = wb.active
        ws.title = "Investors"

        title = f"Investors Export – Generated {datetime.datetime.now():%Y-%m-%d %H:%M}"
        ws.merge_cells('A1:I1')
        ws['A1'] = title
        ws['A1'].font = ws['A1'].font.copy(bold=True, size=14)

        headers = [
            'First Name',
            'Last Name',
            'Phone Number',
            'Email',
            'Region',
            'Currency',
            'Investment Period',
            'Funds Available',
            'Created By',
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = cell.font.copy(bold=True)

        for row_idx, investor in enumerate(qs, start=3):
            row = [
                investor.first_name,
                investor.last_name,
                str(investor.phone_number),
                investor.email if investor.email else 'N/A',
                investor.get_region_display() if investor.region else 'N/A',
                investor.currency,
                investor.investment_period,
                str(investor.funds_available),
                str(investor.created_by),
            ]
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        for i, _ in enumerate(headers, start=1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].auto_size = True

        filename = f"investors_{datetime.date.today():%Y-%m-%d}.xlsx"
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response