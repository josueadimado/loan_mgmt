from django.shortcuts import render, get_object_or_404
from django.urls import reverse_lazy
from django.views.generic import (
    ListView, CreateView, UpdateView, DetailView, View
)
from django.http import HttpResponse
from django.contrib import messages
import csv
from datetime import datetime
from decimal import Decimal
from .models import Borrower, BorrowerDocument
from .forms import BorrowerForm
from loans.models import Loan

class BorrowerListView(ListView):
    model = Borrower
    template_name = 'borrowers/borrower_list.html'
    context_object_name = 'borrowers'
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['count'] = Borrower.objects.count()
        ctx['form'] = BorrowerForm()
        return ctx

class BorrowerCreateView(CreateView):
    model = Borrower
    form_class = BorrowerForm
    template_name = 'borrowers/borrower_form.html'
    success_url = reverse_lazy('borrowers:borrower-list')
    def form_valid(self, form):
        response = super().form_valid(form)
        return response

class BorrowerUpdateView(UpdateView):
    model = Borrower
    form_class = BorrowerForm
    template_name = 'borrowers/borrower_form.html'
    success_url = reverse_lazy('borrowers:borrower-list')
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['borrower'] = self.object
        return context
    def form_valid(self, form):
        changed = False
        changes = []
        for field in form.fields:
            if field in ['id_doc_file', 'guarantor_id_file', 'additional_docs']:
                continue
            submitted_value = form.cleaned_data.get(field)
            existing_value = getattr(self.object, field)
            if field in ['phone_number', 'guarantor_phone']:
                submitted_value_str = str(submitted_value).replace(' ', '') if submitted_value else ''
                existing_value_str = str(existing_value).replace(' ', '') if existing_value else ''
                if submitted_value_str != existing_value_str:
                    changes.append(f"{field}: submitted={submitted_value_str}, existing={existing_value_str}")
                    changed = True
            else:
                if submitted_value != existing_value:
                    if field == 'region':
                        submitted_display = form.fields[field].choices[int(submitted_value)][1] if submitted_value else None
                        existing_display = self.object.get_region_display() if existing_value else None
                        if submitted_display != existing_display:
                            changes.append(f"{field}: submitted={submitted_value} ({submitted_display}), existing={existing_value} ({existing_display})")
                            changed = True
                    elif field in ['id_doc_expiry', 'guarantor_id_expiry']:
                        submitted_str = str(submitted_value) if submitted_value else ''
                        existing_str = str(existing_value) if existing_value else ''
                        if submitted_str != existing_str:
                            changes.append(f"{field}: submitted={submitted_value}, existing={existing_value}")
                            changed = True
                    else:
                        changes.append(f"{field}: submitted={submitted_value}, existing={existing_value}")
                        changed = True
        for field in ['id_doc_file', 'guarantor_id_file']:
            submitted_file = form.cleaned_data.get(field)
            existing_file = getattr(self.object, field)
            if field in form.files:
                changes.append(f"New file uploaded in {field}")
                changed = True
            elif submitted_file != existing_file:
                changes.append(f"{field}: submitted={submitted_file}, existing={existing_file}")
                changed = True
        if 'additional_docs' in form.files and form.files.getlist('additional_docs'):
            changes.append("New files uploaded in additional_docs")
            changed = True
        if changes:
            print("Changes detected:", changes)
        else:
            print("No changes detected.")
        if not changed:
            messages.info(self.request, "Nothing changed.")
            return super().form_valid(form)
        response = super().form_valid(form)
        messages.success(self.request, "Borrower updated successfully.")
        return response
    def form_invalid(self, form):
        print("Form is invalid. Errors:", form.errors)
        print("POST data:", self.request.POST)
        return super().form_invalid(form)

class BorrowerDetailView(DetailView):
    model = Borrower
    template_name = 'borrowers/borrower_detail.html'
    context_object_name = 'borrower'
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['extra_docs'] = BorrowerDocument.objects.filter(borrower=self.object)
        return ctx

class BorrowerExportView(View):
    def get(self, request, *args, **kwargs):
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        import datetime
        qs = Borrower.objects.all().order_by('last_name', 'first_name')
        wb = Workbook()
        ws = wb.active
        ws.title = "Borrowers"
        title = f"Borrowers Export – Generated {datetime.datetime.now():%Y-%m-%d %H:%M}"
        ws.merge_cells('A1:G1')
        ws['A1'] = title
        ws['A1'].font = ws['A1'].font.copy(bold=True, size=14)
        headers = [
            'First Name', 'Last Name',
            'Phone Number', 'Email',
            'Address', 'Region',
            'Created At'
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = cell.font.copy(bold=True)
        for row_idx, b in enumerate(qs, start=3):
            row = [
                b.first_name,
                b.last_name,
                str(b.phone_number),
                b.email,
                b.address,
                b.get_region_display(),
                b.created_at.strftime('%Y-%m-%d %H:%M'),
            ]
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        for i, _ in enumerate(headers, start=1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].auto_size = True
        filename = f"borrowers_{datetime.date.today():%Y-%m-%d}.xlsx"
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

class BorrowerStatementView(View):
    template_name = 'borrowers/borrower_statement.html'
    def get(self, request, *args, **kwargs):
        # Get the borrower
        borrower = get_object_or_404(Borrower, pk=self.kwargs['pk'])
        # Fetch all loans for the borrower
        loans = Loan.objects.filter(borrower=borrower).prefetch_related('repayments')
        # Calculate loan summaries and totals
        loan_data = []
        total_principal = Decimal('0.00')
        total_interest = Decimal('0.00')
        total_to_pay = Decimal('0.00')
        total_paid = Decimal('0.00')
        total_remaining = Decimal('0.00')
        for loan in loans:
            total_to_pay_loan = loan.get_total_to_pay()
            total_paid_loan = loan.get_total_paid()
            remaining_loan = loan.get_remaining_amount_to_pay()
            interest_loan = loan.get_total_interest()
            total_principal += loan.principal
            total_interest += interest_loan
            total_to_pay += total_to_pay_loan
            total_paid += total_paid_loan
            total_remaining += remaining_loan
            # Calculate running balances for each repayment
            repayments_with_balances = []
            balance = Decimal(loan.principal)
            for repayment in loan.repayments.all():
                balance -= Decimal(repayment.amount)
                repayments_with_balances.append({
                    'repayment': repayment,
                    'balance': balance,
                })
            loan_data.append({
                'loan': loan,
                'principal': loan.principal,
                'interest_rate': loan.interest_rate,
                'total_to_pay': total_to_pay_loan,
                'total_paid': total_paid_loan,
                'remaining': remaining_loan,
                'start_date': loan.start_date,
                'repayments_with_balances': repayments_with_balances, # Include repayments with balances
            })
        # Prepare context
        context = {
            'borrower': borrower,
            'loan_data': loan_data,
            'total_principal': total_principal,
            'total_interest': total_interest,
            'total_to_pay': total_to_pay,
            'total_paid': total_paid,
            'total_remaining': total_remaining,
            'statement_date': datetime.now(),
            'generated_by': request.user if request.user.is_authenticated else None,
            'auto_print': 'print' in request.GET, # Flag to trigger printing
        }
        # Check if the request is for a PDF download
        if 'download' in request.GET:
            return self.render_to_pdf(context)
        return render(request, self.template_name, context)
    def render_to_pdf(self, context):
        from weasyprint import HTML
        from django.template.loader import render_to_string
        import tempfile
        # Render the HTML template to a string
        html_string = render_to_string(self.template_name, context)
        # Create a PDF
        html = HTML(string=html_string, base_url=self.request.build_absolute_uri())
        pdf_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        html.write_pdf(pdf_file.name)
        # Serve the PDF as a response
        with open(pdf_file.name, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/pdf')
            filename = f"statement_{context['borrower'].first_name}_{context['borrower'].last_name}_{datetime.now().strftime('%Y%m%d')}.pdf"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
        pdf_file.close()
        import os
        os.unlink(pdf_file.name)
        return response